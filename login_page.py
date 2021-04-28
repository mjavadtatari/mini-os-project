import sys
import openpyxl
from PyQt5 import QtCore
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel, QLineEdit, QGridLayout, QMessageBox)


class ChangePassword(QWidget):
    """docstring for ChangePassword."""

    def __init__(self):
        super(ChangePassword, self).__init__()

        self.setWindowTitle('miniOS')
        self.resize(500, 256)

        layout = QGridLayout()

        label_name = QLabel(
            '<font size="4"><center> Change Your Password </center></font>')
        layout.addWidget(label_name, 0, 0, 1, 2)

        label_password_1 = QLabel('<font size="4"> Password </font>')
        self.lineEdit_password_1 = QLineEdit()
        self.lineEdit_password_1.setEchoMode(2)
        self.lineEdit_password_1.setPlaceholderText(
            'Please enter your password')
        layout.addWidget(label_password_1, 2, 0)
        layout.addWidget(self.lineEdit_password_1, 2, 1)

        label_password_2 = QLabel('<font size="4"> Password </font>')
        self.lineEdit_password_2 = QLineEdit()
        self.lineEdit_password_2.setEchoMode(2)
        self.lineEdit_password_2.setPlaceholderText(
            'Please Re-enter your password')
        layout.addWidget(label_password_2, 2, 0)
        layout.addWidget(self.lineEdit_password_2, 2, 1)

        button_submit = QPushButton('Submit')
        button_submit.clicked.connect(self.save_password)
        layout.addWidget(button_submit, 3, 0, 1, 2)
        layout.setRowMinimumHeight(2, 25)

        self.setLayout(layout)

    def save_password(self):
        users_db = openpyxl.load_workbook('Users.xlsx')
        users_db_ws = users_db.active
        msg = QMessageBox()
        founded = False

        # for row in users_db_ws.iter_rows(users_db_ws.min_row, users_db_ws.max_row + 1):
        #     for cell in row:
        #         if str(cell.value) == self.lineEdit_username.text():
        #             if str(users_db_ws.cell(row=cell.row, column=2).value) == self.lineEdit_password.text():
        #                 founded = True
        #                 msg.setText('Success')
        #                 msg.exec_()
        #                 chw=ChangePassword()
        #                 chw.show()
        #                 # app.quit()
        if not founded:
            msg.setText('Incorrect Username or Password')
            msg.exec_()


class LoginForm(QWidget):

    switch_window = QtCore.pyqtSignal()
    loggedInUsername = ''

    def __init__(self, *args, **kwargs):
        super(LoginForm, self).__init__(*args, **kwargs)
        # self.setWindowTitle('miniOS')
        # self.resize(500, 256)

        layout = QGridLayout()

        label_name = QLabel(
            '<font size="4"><h2><center> miniOS </center></h2></font>')
        layout.addWidget(label_name, 0, 0, 1, 2)

        label_message = QLabel('')
        layout.addWidget(label_message, 1, 0)

        label_name = QLabel('<font size="4"> Username </font>')
        self.lineEdit_username = QLineEdit()
        self.lineEdit_username.setPlaceholderText('Please enter your username')
        layout.addWidget(label_name, 2, 0)
        layout.addWidget(self.lineEdit_username, 2, 1)

        label_password = QLabel('<font size="4"> Password </font>')
        self.lineEdit_password = QLineEdit()
        self.lineEdit_password.setEchoMode(2)
        self.lineEdit_password.setPlaceholderText('Please enter your password')
        layout.addWidget(label_password, 3, 0)
        layout.addWidget(self.lineEdit_password, 3, 1)

        button_login = QPushButton('Login')
        button_login.clicked.connect(self.check_password)
        layout.addWidget(button_login, 4, 0, 1, 2)
        layout.setRowMinimumHeight(2, 25)

        self.setLayout(layout)

    def check_password(self):
        users_db = openpyxl.load_workbook('Users.xlsx')
        users_db_ws = users_db.active
        # msg = QMessageBox()
        founded = False

        for row in users_db_ws.iter_rows(users_db_ws.min_row, users_db_ws.max_row + 1):
            for cell in row:
                if str(cell.value) == self.lineEdit_username.text():
                    if str(users_db_ws.cell(row=cell.row, column=2).value) == self.lineEdit_password.text():
                        founded = True
                        self.loggedInUsername = self.lineEdit_username.text()
                        self.connect(self.login)
                        # msg.setText('Success')
                        # msg.exec_()
                        # chw=ChangePassword()
                        # chw.show()
                        # app.quit()
        if not founded:
            label_message = QLabel('Incorrect Username or Password')
            # layout.addWidget(label_message, 1, 0)
            # msg.setText('Incorrect Username or Password')
            # msg.exec_()

    def get_loggedInUsername(self):
        return self.loggedInUsername

    def login(self):
        self.switch_window.emit()


if __name__ == '__main__':
    app = QApplication(sys.argv)

    form = LoginForm()
    form.show()

    sys.exit(app.exec_())
