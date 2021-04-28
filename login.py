import sys
import openpyxl
from management import *
from logs_file import *
from termcolor import colored


users_db = openpyxl.load_workbook('Users.xlsx')
users_db_ws = users_db.active


class LoggedInUser():

    def __init__(self, *args, **kwargs):
        self.username = ''
        self.password = ''
        self.first_name = ''
        self.last_name = ''
        self.email = ''
        self.failed_attempted = 0
        self.user_row = ''
        self.logged_in = False

    def find_user_coordinate(self, username):
        for row in users_db_ws.iter_rows(users_db_ws.min_row, users_db_ws.max_row + 1):
            for cell in row:
                if str(cell.value) == username:
                    self.user_row = str(cell.row)
                    self.load_user_info()

        return False if self.user_row == '' else True

    def load_user_info(self):
        self.username = users_db_ws['A' + self.user_row].value
        self.password = users_db_ws['B' + self.user_row].value
        self.email = users_db_ws['D' + self.user_row].value
        self.first_name = users_db_ws['E' + self.user_row].value
        self.last_name = users_db_ws['F' + self.user_row].value

    def check_password(self, username, password):
        if self.find_user_coordinate(username) and self.password == password:
            if not self.check_is_user_banned():
                self.unban_user()
                self.logged_in = True
                add_record(self.username, 'login', 'pass: ' +
                           self.password, 'Success')
                return True
            else:
                add_record(self.username, 'login', 'pass: ' +
                           self.password, 'Failed, The User is Banned')
                print(colored(
                    'Your account has been banned, please contact the administrator!', 'red'))
                sleep(3)
                quit()
        else:
            return False

    def check_is_user_banned(self):
        if users_db_ws['H' + self.user_row].value == 'T':
            return True
        else:
            return False

    def ban_user(self, username, password):
        if username == 'mjavadtatri':
            return True

        print('asda' + self.username)

        if self.failed_attempted >= 2 and self.username != '':
            users_db_ws['H' + self.user_row].value = 'T'
            add_record(username, 'banned', '3 failed attempts',
                       'Banned Successfully')
            print(
                colored('due to 3 failed attempts, your account has been banned!', 'red'))
            sleep(3)
            quit()

        elif self.failed_attempted < 2:
            self.failed_attempted += 1
            add_record(username, 'login', 'pass: ' + password,
                       'Failed Attempts= ' + str(self.failed_attempted))
            print(colored('your failed attempts= ' +
                  str(self.failed_attempted) + '\n\n\n', 'yellow'))

        else:
            add_record(username, 'kicked', '3 failed attempts',
                       'Session has Ended')
            print(
                colored('due to 3 failed attempts, your session has ended!', 'red'))
            sleep(3)
            quit()

    def unban_user(self):
        self.failed_attempted = 0
        users_db_ws['H' + self.user_row].value = 'F'

    def loginPage(self):
        while True:
            temp_username = input('Login as: ')
            temp_password = input(temp_username + '\'s Password: ')

            if self.check_password(temp_username, temp_password):
                print(colored('\nWelcome ' + temp_username + '\n\n\n', 'green'))
                break
            else:
                print(colored('\nThe Username or Password is Incorrect!', 'red'))
                self.ban_user(temp_username, temp_password)
