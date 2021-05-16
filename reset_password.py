import smtplib
import ssl
import random
import sys
import time
import openpyxl
from management import *
from logs_file import *

users_db = openpyxl.load_workbook('Users.xlsx')
users_db_ws = users_db.active


def forget_password(temp_username):
    for row in users_db_ws.iter_rows(users_db_ws.min_row, users_db_ws.max_row + 1):
        for cell in row:
            if str(cell.value) == temp_username:
                temp_email = users_db_ws['D' + str(cell.row)].value
                if not users_db_ws['H' + str(cell.row)].value == 'T':
                    if temp_email:
                        add_record(temp_username, 'ForgetPass',
                                   'Email: {}'.format(temp_email), 'Success, Email Found!')
                        send_gmail(temp_email, temp_username)
                    else:
                        add_record(temp_username, 'ForgetPass',
                                   'Email: {}'.format(temp_email), 'Failed, Email Not Found!')
                        print(colored('Email Not Found!', 'red'))
                        time.sleep(3)
                        sys.exit()
                else:
                    add_record(temp_username, 'ForgetPass',
                               'Email: {}'.format(temp_email), 'Failed, User is Banned!')
                    print(colored('Your Account Has Been Banned!', 'red'))
                    time.sleep(3)
                    sys.exit()


def set_new_password(temp_username):
    characters = "01234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ&*!@#$%^abcdefghijklmnopqrstuvwxyz"
    # Take the length of the password from the user
    password_length = 16
    # Generate the password
    password = "".join(random.sample(characters, password_length))
    # Print the generated password

    for row in users_db_ws.iter_rows(users_db_ws.min_row, users_db_ws.max_row + 1):
        for cell in row:
            if str(cell.value) == temp_username:
                users_db_ws['C' + str(cell.row)
                            ].value = users_db_ws['B' + str(cell.row)].value
                users_db_ws['B' + str(cell.row)].value = password
                users_db.save('Users.xlsx')

    add_record(temp_username, 'SetNewPass', 'New Pass: {}'.format(password),
               'Success. Password Changed!')
    return password


def send_gmail(receiver_email, username):
    print(colored('Sending Email, Please Wait and Do NOT Close the Program!', 'yellow'))
    new_password = set_new_password(username)

    smtp_server = "smtp.gmail.com"
    port = 587  # For starttls
    sender_email = "minios.info@gmail.com"
    password = '7ESy1L*Dyoj2FhNHha6&'

    subject = 'ReNew Password!'
    text = """
    Hi {},
    Your New Password: {}

    This message is sent from miniOS.""".format(username, new_password)

    message = 'Subject: {}\n\n{}'.format(subject, text)

    # Create a secure SSL context
    context = ssl.create_default_context()

    # Try to log in to server and send email
    with smtplib.SMTP(smtp_server, port) as server:
        server.ehlo()  # Can be omitted
        server.starttls(context=context)
        server.ehlo()  # Can be omitted
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message)

    print(colored('Email Sent Successfully!', 'green'))
    add_record(username, 'SendEmail', 'New Pass: {}'.format(new_password),
               'Success. Email Sent to {}'.format(receiver_email))

    print(colored('The Program Closes Automatically!', 'yellow'))
    time.sleep(3)
    sys.exit()
