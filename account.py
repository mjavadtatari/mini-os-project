import sys
import openpyxl
from management import *
from logs_file import *
# from termcolor import colored
from reset_password import forget_password


users_db = openpyxl.load_workbook('Users.xlsx')
users_db_ws = users_db.active


class Account():

    def __init__(self, *args, **kwargs):
        self.username = ''
        self.password = ''
        self.first_name = ''
        self.last_name = ''
        self.email = ''
        self.failed_attempted = 0
        self.user_row = ''
        self.logged_in = False
        self.relogin = False

    def find_user_coordinate(self, username):
        for row in users_db_ws.iter_rows(users_db_ws.min_row, users_db_ws.max_row + 1):
            for cell in row:
                if str(cell.value) == username:
                    self.user_row = str(cell.row)
                    self.load_user_info()

        return False if self.user_row == '' else True

    def load_user_info(self):
        self.username = users_db_ws['A' + self.user_row].value
        self.password = str(users_db_ws['B' + self.user_row].value)
        self.email = users_db_ws['D' + self.user_row].value
        self.first_name = users_db_ws['E' + self.user_row].value
        self.last_name = users_db_ws['F' + self.user_row].value

    def check_password(self, username, password):
        if self.find_user_coordinate(username) and self.password == password:
            if not self.check_is_user_banned():
                self.unban_user()
                self.logged_in = True
                add_record(self.username, 'login', 'pass: ' +
                           self.password, 'Success')
                return True
            else:
                add_record(self.username, 'login', 'pass: ' +
                           self.password, 'Failed, The User is Banned', 'r')
                print('Your account has been banned, please contact the administrator!')
                sleep(3)
                quit()
        else:
            return False

    def check_is_user_banned(self, temp_username=None):
        if temp_username:
            for row in users_db_ws.iter_rows(users_db_ws.min_row, users_db_ws.max_row + 1):
                for cell in row:
                    if str(cell.value) == temp_username:
                        temp_user_row = str(cell.row)

            if users_db_ws['H' + temp_user_row].value == 'T':
                return True
            else:
                return False

    def ban_user(self, username, password):
        if username == 'mjavadtatri':
            return True

        if self.failed_attempted >= 3 and self.username != '':
            users_db_ws['H' + self.user_row].value = 'T'
            users_db.save('Users.xlsx')
            add_record(username, 'banned', '3 failed attempts',
                       'User Banned Successfully!')
            print('due to 3 failed attempts, your account has been banned!')
            sleep(3)
            quit()

        elif self.failed_attempted < 3:
            self.failed_attempted += 1
            add_record(username, 'login', 'pass: ' + password,
                       'Failed Attempts= ' + str(self.failed_attempted), 'r')
            print('your failed attempts= ' +
                  str(self.failed_attempted) + '\n\n\n')

        else:
            add_record(username, 'kicked', '3 failed attempts',
                       'Session has Ended', 'r')
            print('due to 3 failed attempts, your session has ended!')
            sleep(3)
            quit()

    def unban_user(self):
        self.failed_attempted = 0
        users_db_ws['H' + self.user_row].value = 'F'
        users_db.save('Users.xlsx')

    def set_email(self):
        import re
        regex = '^(\w|\.|\_|\-)+[@](\w|\_|\-|\.)+[.]\w{2,3}$'

        if str(users_db_ws['D' + self.user_row].value) == 'None':
            print('\n\nYou Need to Submit Your Email Address!\n')
            while True:
                temp_email = input(self.username_title() + 'Email : ')
                if re.search(regex, temp_email):
                    users_db_ws['D' + self.user_row].value = temp_email
                    users_db.save('Users.xlsx')
                    add_record(self.username, 'set email', 'email :' + temp_email,
                               'Submited Successfully')
                    print('\nEmail Submited Successfully!\n\n')
                    break
                else:
                    add_record(self.username, 'set email', 'email :' + temp_email,
                               'Failed, Not a Valid address', 'r')
                    print('Enter a Valid Email address!\n')

    def password_strength_checker(self, password):
        import re

        if len(password) < 8:
            return False

        if not re.search("[a-z]", password):
            return False

        if not re.search("[A-Z]", password):
            return False

        if not re.search("[$#@^*%&]", password):
            return False

        if re.search("\s", password):
            return False

        return True

    def change_password(self):
        self.set_email()
        if users_db_ws['G' + self.user_row].value == 'T':
            print('\nYou Need to Change Your Password!')
            print('The password must be at least 8 characters long and contain both lower and upper case letters, numbers, and symbols.\n\n')
            while True:
                temp_password = input(
                    self.username_title() + 'New Password : ')
                if self.password_strength_checker(temp_password):
                    users_db_ws['C' + self.user_row].value = users_db_ws['B' +
                                                                         self.user_row].value
                    users_db_ws['B' + self.user_row].value = temp_password
                    users_db_ws['G' + self.user_row].value = 'F'
                    users_db.save('Users.xlsx')
                    self.relogin = True
                    add_record(self.username, 'change password', 'pass :' + temp_password,
                               'Changed Successfully')
                    print('\nThe new password is Great!, Login again!!\n\n\n')
                    break
                else:
                    print('The entered Password is not Strong enough, Try again!\n')
                    add_record(self.username, 'change password', 'pass :' + temp_password,
                               'Failed, easily crackable password', 'r')
        else:
            return True

    def loginPage(self):
        while True:
            if self.failed_attempted >= 3 and self.username != '':
                users_db_ws['H' + self.user_row].value = 'T'
                users_db.save('Users.xlsx')
                add_record(self.username, 'banned', '3 failed attempts',
                           'User Banned Successfully!')
                print('due to 3 failed attempts, your account has been banned!')
                sleep(3)
                quit()

            temp_username = input('Login as: ')

            if self.check_is_user_banned(temp_username):
                add_record(temp_username, 'kicked', 'User is Banned',
                           'Success, Session has Ended')
                print('Your Account Has Been Banned!')
                sleep(3)
                quit()

            temp_password = input(
                temp_username + '\'s Password (\'F\' for forgot Password): ')

            if temp_password.lower() == 'f':
                forget_password(temp_username)
                break

            if self.check_password(temp_username, temp_password):
                print('\nWelcome ' + temp_username)
                self.change_password()
                break
            else:
                print('\nThe Username or Password is Incorrect!')
                self.ban_user(temp_username, temp_password)

    def username_title(self):
        return '{} >> '.format(self.username)

    def update_account_info(self):
        temp_first = input(self.username_title() + 'First Name : ')
        temp_last = input(self.username_title() + 'Last Name : ')
