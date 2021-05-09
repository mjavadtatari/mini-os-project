import openpyxl
import os
import datetime
from openpyxl.styles import Font
from termcolor import colored

logs_file_dir = 'root\\administrator'
logs_file = None


def open_logs_file():
    # Try to open or create logs_file.xlsx
    global logs_file

    if not os.path.exists('root'):
        # Try to open or create directory.
        os.makedirs('root')

    if not os.path.exists(logs_file_dir):
        # Try to open or create directory.
        os.makedirs(logs_file_dir)

    if os.path.isfile(logs_file_dir + '\\logs_file.xlsx'):
        # Try to find the xlsx file or create a new one.
        # Opens and returns the worksheet.
        logs_file = openpyxl.load_workbook(logs_file_dir + '\\logs_file.xlsx')
        logs_file_worksheet = logs_file['system logs']
    else:
        logs_file = openpyxl.Workbook()
        logs_file_worksheet = logs_file.active

        logs_file_worksheet.title = 'system logs'

        logs_file_worksheet['A1'] = 'Date'
        logs_file_worksheet['B1'] = 'Time'
        logs_file_worksheet['C1'] = 'Username'
        logs_file_worksheet['D1'] = 'Command'
        logs_file_worksheet['E1'] = 'Input Parameters'
        logs_file_worksheet['F1'] = 'Output Resault'

        logs_file_worksheet.column_dimensions['A'].width = 14
        logs_file_worksheet.column_dimensions['B'].width = 10
        logs_file_worksheet.column_dimensions['C'].width = 15
        logs_file_worksheet.column_dimensions['D'].width = 20
        logs_file_worksheet.column_dimensions['E'].width = 85
        logs_file_worksheet.column_dimensions['F'].width = 40

        for rows in logs_file_worksheet.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color='FF3498DB', end_color='FF3498DB', fill_type="solid")

    return logs_file_worksheet


def save_logs_file():
    # Saving the xlsx as logs_file.
    global logs_file

    logs_file.save(logs_file_dir + '\\logs_file.xlsx')

    return True


def find_last_line():
    ws = open_logs_file()
    return ws, str(ws.max_row + 1)


def now_date_time():
    now = datetime.datetime.now()
    return now.strftime("%Y/%m/%d"), now.strftime("%H:%M:%S")


def add_record(r_user, r_command, r_input, r_output, r_color=None):
    # Adding a record to the xlsx file.
    ws, last_line = find_last_line()

    ws['A' + last_line], ws['B' + last_line] = now_date_time()
    ws['C' + last_line] = r_user
    ws['D' + last_line] = str(r_command)
    ws['E' + last_line] = str(r_input)
    ws['F' + last_line] = str(r_output)
    if 'fail' in r_output.lower():
        ws['F' + last_line].font = Font(color='FFE74C3C')

    elif 'success' in r_output.lower():
        ws['F' + last_line].font = Font(color='FF27AE60')

    save_logs_file()
